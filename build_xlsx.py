"""
build_xlsx.py
Builds capstone_spreadsheet.xlsx as an exact mirror of the Google Sheet built
for Part 2 (Monthly Data, Category Targets, Pivot, Category Summary), with
live formulas and conditional formatting, so opening it in Excel/LibreOffice
reproduces the same verified numbers as the live Google Sheet.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

wb = Workbook()

# ------------------------------------------------------------------
# Sheet 1: Monthly Data (verbatim copy of monthly_category_revenue.csv)
# ------------------------------------------------------------------
ws1 = wb.active
ws1.title = "Monthly Data"

with open("monthly_category_revenue.csv", newline="") as f:
    reader = csv.reader(f)
    rows = list(reader)

for r_idx, row in enumerate(rows, start=1):
    for c_idx, val in enumerate(row, start=1):
        if r_idx == 1:
            ws1.cell(row=r_idx, column=c_idx, value=val)
        else:
            # numeric columns: order_count (3), total_revenue (4), avg_revenue (5)
            if c_idx in (3, 4, 5):
                ws1.cell(row=r_idx, column=c_idx, value=float(val) if "." in val else int(val))
            else:
                ws1.cell(row=r_idx, column=c_idx, value=val)

header_font = Font(bold=True)
for c_idx in range(1, 6):
    ws1.cell(row=1, column=c_idx).font = header_font
for col, width in zip("ABCDE", (22, 12, 13, 15, 14)):
    ws1.column_dimensions[col].width = width

n_data_rows = len(rows) - 1  # 36

# ------------------------------------------------------------------
# Sheet 2: Category Targets
# ------------------------------------------------------------------
ws2 = wb.create_sheet("Category Targets")
targets = [
    ("Fruits & Vegetables", 12000),
    ("Dairy & Eggs", 16500),
    ("Snacks & Beverages", 13000),
    ("Personal Care", 15500),
    ("Household Essentials", 17000),
    ("Bakery", 12000),
]
ws2.append(["category", "target_revenue_inr"])
for cat, tgt in targets:
    ws2.append([cat, tgt])
for c_idx in range(1, 3):
    ws2.cell(row=1, column=c_idx).font = header_font
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 18

# ------------------------------------------------------------------
# Sheet 3: Pivot (category-level SUM of total_revenue / order_count,
# computed with SUMIF formulas against Monthly Data — a live pivot
# equivalent, not a hardcoded snapshot)
# ------------------------------------------------------------------
ws3 = wb.create_sheet("Pivot")
pivot_categories = [
    "Bakery",
    "Dairy & Eggs",
    "Fruits & Vegetables",
    "Household Essentials",
    "Personal Care",
    "Snacks & Beverages",
]
ws3.append(["category", "SUM of total_revenue", "SUM of order_count"])
last_md_row = n_data_rows + 1  # row 37
for i, cat in enumerate(pivot_categories, start=2):
    ws3.cell(row=i, column=1, value=cat)
    ws3.cell(
        row=i, column=2,
        value=f"=SUMIF('Monthly Data'!A2:A{last_md_row},A{i},'Monthly Data'!D2:D{last_md_row})",
    )
    ws3.cell(
        row=i, column=3,
        value=f"=SUMIF('Monthly Data'!A2:A{last_md_row},A{i},'Monthly Data'!C2:C{last_md_row})",
    )
grand_row = len(pivot_categories) + 2  # row 8
ws3.cell(row=grand_row, column=1, value="Grand Total").font = header_font
ws3.cell(row=grand_row, column=2, value=f"=SUM(B2:B{grand_row - 1})").font = header_font
ws3.cell(row=grand_row, column=3, value=f"=SUM(C2:C{grand_row - 1})").font = header_font
for c_idx in range(1, 4):
    ws3.cell(row=1, column=c_idx).font = header_font
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 20

# ------------------------------------------------------------------
# Sheet 4: Category Summary
# ------------------------------------------------------------------
ws4 = wb.create_sheet("Category Summary")
headers = [
    "category", "total_revenue", "order_count", "target_revenue_inr",
    "variance", "percentage_variance", "target_status",
    "sql_reference_total", "matches_part1_sql_total",
]
ws4.append(headers)
for c_idx in range(1, len(headers) + 1):
    ws4.cell(row=1, column=c_idx).font = header_font

sql_reference = {
    "Bakery": 15410,
    "Dairy & Eggs": 14090,
    "Fruits & Vegetables": 9790,
    "Household Essentials": 21715,
    "Personal Care": 16382,
    "Snacks & Beverages": 10895,
}

for i, cat in enumerate(pivot_categories, start=2):
    ws4.cell(row=i, column=1, value=cat)
    ws4.cell(row=i, column=2, value=f"=VLOOKUP(A{i},Pivot!A:C,2,FALSE)")
    ws4.cell(row=i, column=3, value=f"=VLOOKUP(A{i},Pivot!A:C,3,FALSE)")
    ws4.cell(
        row=i, column=4,
        value=(
            f"=IFERROR(XLOOKUP(A{i},'Category Targets'!A:A,'Category Targets'!B:B),"
            f"IFERROR(VLOOKUP(A{i},'Category Targets'!A:B,2,FALSE),\"Not Found\"))"
        ),
    )
    ws4.cell(row=i, column=5, value=f"=D{i}-B{i}")
    ws4.cell(row=i, column=6, value=f"=((B{i}-D{i})*100)/D{i}")
    ws4.cell(
        row=i, column=7,
        value=(
            f'=IF(B{i}>=D{i},"Above Target",'
            f'IF(((D{i}-B{i})*100/D{i})<=15,"Below Target - Watch",'
            f'"Below Target - Critical"))'
        ),
    )
    ws4.cell(row=i, column=8, value=sql_reference[cat])
    ws4.cell(row=i, column=9, value=f'=IF(B{i}=H{i},"Yes","No")')

for col, width in zip("ABCDEFGHI", (22, 14, 12, 18, 10, 18, 22, 18, 22)):
    ws4.column_dimensions[col].width = width

last_row = len(pivot_categories) + 1  # row 7
status_range = f"G2:G{last_row}"

green_fill = PatternFill(start_color="B7D7A8", end_color="B7D7A8", fill_type="solid")
amber_fill = PatternFill(start_color="F9CB9C", end_color="F9CB9C", fill_type="solid")
red_fill = PatternFill(start_color="EA9999", end_color="EA9999", fill_type="solid")

ws4.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"Above Target"'], fill=green_fill),
)
ws4.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"Below Target - Watch"'], fill=amber_fill),
)
ws4.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"Below Target - Critical"'], fill=red_fill),
)

wb.save("capstone_spreadsheet.xlsx")
print("Saved capstone_spreadsheet.xlsx")
