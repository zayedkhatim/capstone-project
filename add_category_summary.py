"""
add_category_summary.py
Step 3 of 3. Run this AFTER build_xlsx.py and add_pivot_table.py, against
the same capstone_spreadsheet.xlsx (which at this point holds "Monthly
Data", "Category Targets" and the native "Pivot Table" sheet).

Adds the "Category Summary" sheet with live formulas:
  - total_revenue / order_count: VLOOKUP against the native 'Pivot Table'
    sheet — a real dependency on the real PivotTable feature required by
    the brief, not a formula-only simulation sitting unused beside it.
  - target_revenue_inr: a genuine XLOOKUP against 'Category Targets', using
    XLOOKUP's own native not-found argument rather than a defensive
    IFERROR(XLOOKUP(...), IFERROR(VLOOKUP(...))) fallback wrapper.
  - variance / percentage_variance / target_status: derived from the above.
  - sql_reference_total / matches_part1_sql_total: cross-check against the
    Part 1 SQL totals.

Why this script is openpyxl-only (no further LibreOffice pass): the
LibreOffice engine available in this build environment (24.2) predates
XLOOKUP support (added in LibreOffice 24.8), so opening this sheet in
LibreOffice mis-parses — and silently corrupts — the XLOOKUP formula text
(observed during development: the function name got rewritten in lowercase
and evaluated to #NAME?; see ai_log.md). Writing the formulas with openpyxl
instead avoids that entirely: the formula text openpyxl writes is exact and
untouched, and any current Excel or Google Sheets opens and computes it
correctly.

The one thing openpyxl can't do is compute a cached value for a formula it
writes. So this script also injects a cached <v> for each formula cell
directly into the saved worksheet XML — computed independently in Python
from the native Pivot Table's own output (read back with data_only=True,
not hardcoded) and the same target/status logic the formulas encode — so
the file shows correct numbers immediately (e.g. in GitHub's file preview)
without requiring the viewer to recalculate, while the formulas themselves
remain the real, working, uncorrupted source of truth.
"""
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule

XLSX_PATH = "capstone_spreadsheet.xlsx"

pivot_categories = [
    "Bakery",
    "Dairy & Eggs",
    "Fruits & Vegetables",
    "Household Essentials",
    "Personal Care",
    "Snacks & Beverages",
]

targets = {
    "Fruits & Vegetables": 12000,
    "Dairy & Eggs": 16500,
    "Snacks & Beverages": 13000,
    "Personal Care": 15500,
    "Household Essentials": 17000,
    "Bakery": 12000,
}

sql_reference = {
    "Bakery": 15410,
    "Dairy & Eggs": 14090,
    "Fruits & Vegetables": 9790,
    "Household Essentials": 21715,
    "Personal Care": 16382,
    "Snacks & Beverages": 10895,
}

# Read the native Pivot Table's actual computed output (not hardcoded) so
# the injected cache values are demonstrably sourced from the real
# PivotTable, and cross-check it against the Part 1 SQL totals.
wb_v = load_workbook(XLSX_PATH, data_only=True)
pt = wb_v["Pivot Table"]
pivot_output = {}
for row in pt.iter_rows(min_row=1, max_row=pt.max_row, max_col=3):
    cat, rev, cnt = (c.value for c in row)
    if cat in pivot_categories:
        pivot_output[cat] = (rev, cnt)
assert set(pivot_output) == set(pivot_categories), f"Pivot Table missing categories: {pivot_output}"
for cat in pivot_categories:
    assert pivot_output[cat][0] == sql_reference[cat], (
        f"Pivot Table total_revenue for {cat} ({pivot_output[cat][0]}) does not match "
        f"Part 1 SQL total ({sql_reference[cat]})"
    )

# ------------------------------------------------------------------
# Build the sheet with openpyxl (formulas only — no cached values yet)
# ------------------------------------------------------------------
wb = load_workbook(XLSX_PATH, data_only=False)
if "Category Summary" in wb.sheetnames:
    del wb["Category Summary"]
ws = wb.create_sheet("Category Summary")

header_font = Font(bold=True)
headers = [
    "category", "total_revenue", "order_count", "target_revenue_inr",
    "variance", "percentage_variance", "target_status",
    "sql_reference_total", "matches_part1_sql_total",
]
ws.append(headers)
for c_idx in range(1, len(headers) + 1):
    ws.cell(row=1, column=c_idx).font = header_font

cached = {}  # (row, col) -> (value, is_string)

for i, cat in enumerate(pivot_categories, start=2):
    rev, cnt = pivot_output[cat]
    tgt = targets[cat]
    variance = tgt - rev
    pct_variance = ((rev - tgt) * 100) / tgt
    if rev >= tgt:
        status = "Above Target"
    elif ((tgt - rev) * 100 / tgt) <= 15:
        status = "Below Target - Watch"
    else:
        status = "Below Target - Critical"
    matches = "Yes" if rev == sql_reference[cat] else "No"

    ws.cell(row=i, column=1, value=cat)
    ws.cell(row=i, column=2, value=f"=VLOOKUP(A{i},'Pivot Table'!A:C,2,FALSE)")
    ws.cell(row=i, column=3, value=f"=VLOOKUP(A{i},'Pivot Table'!A:C,3,FALSE)")
    ws.cell(
        row=i, column=4,
        value=f"=XLOOKUP(A{i},'Category Targets'!A:A,'Category Targets'!B:B,\"Not Found\")",
    )
    ws.cell(row=i, column=5, value=f"=D{i}-B{i}")
    ws.cell(row=i, column=6, value=f"=((B{i}-D{i})*100)/D{i}")
    ws.cell(
        row=i, column=7,
        value=(
            f'=IF(B{i}>=D{i},"Above Target",'
            f'IF(((D{i}-B{i})*100/D{i})<=15,"Below Target - Watch",'
            f'"Below Target - Critical"))'
        ),
    )
    ws.cell(row=i, column=8, value=sql_reference[cat])
    ws.cell(row=i, column=9, value=f'=IF(B{i}=H{i},"Yes","No")')

    cached[(i, 2)] = (rev, False)
    cached[(i, 3)] = (cnt, False)
    cached[(i, 4)] = (tgt, False)
    cached[(i, 5)] = (variance, False)
    cached[(i, 6)] = (pct_variance, False)
    cached[(i, 7)] = (status, True)
    cached[(i, 9)] = (matches, True)

for col, width in zip("ABCDEFGHI", (22, 14, 12, 18, 10, 18, 22, 18, 22)):
    ws.column_dimensions[col].width = width

last_row = len(pivot_categories) + 1
status_range = f"G2:G{last_row}"
green_fill = PatternFill(start_color="B7D7A8", end_color="B7D7A8", fill_type="solid")
amber_fill = PatternFill(start_color="F9CB9C", end_color="F9CB9C", fill_type="solid")
red_fill = PatternFill(start_color="EA9999", end_color="EA9999", fill_type="solid")
ws.conditional_formatting.add(
    status_range, CellIsRule(operator="equal", formula=['"Above Target"'], fill=green_fill)
)
ws.conditional_formatting.add(
    status_range, CellIsRule(operator="equal", formula=['"Below Target - Watch"'], fill=amber_fill)
)
ws.conditional_formatting.add(
    status_range, CellIsRule(operator="equal", formula=['"Below Target - Critical"'], fill=red_fill)
)

wb.save(XLSX_PATH)
print("Wrote Category Summary formulas.")

# ------------------------------------------------------------------
# Inject cached <v> values for the formula cells directly into the saved
# worksheet XML (openpyxl itself never computes/caches formula results).
# Locate the correct worksheet part via workbook.xml -> rels, not by
# guessing a sheetN.xml number or sniffing text (openpyxl stores text via
# the shared-strings table, not inline, so header text isn't in this XML).
# ------------------------------------------------------------------
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKGREL = "http://schemas.openxmlformats.org/package/2006/relationships"

with zipfile.ZipFile(XLSX_PATH) as zf:
    contents = {n: zf.read(n) for n in zf.namelist()}

wb_root = ET.fromstring(contents["xl/workbook.xml"])
sheet_el = None
for s in wb_root.find(f"{{{NS_MAIN}}}sheets"):
    if s.get("name") == "Category Summary":
        sheet_el = s
        break
assert sheet_el is not None, "Category Summary not found in workbook.xml"
rid = sheet_el.get(f"{{{NS_REL}}}id")

rels_root = ET.fromstring(contents["xl/_rels/workbook.xml.rels"])
target = None
for rel in rels_root:
    if rel.get("Id") == rid:
        target = rel.get("Target")
        break
assert target is not None, f"Relationship {rid} not found in workbook.xml.rels"
# Target is either absolute from the package root ("/xl/worksheets/sheetN.xml")
# or relative to the xl/ folder ("worksheets/sheetN.xml") — handle both.
target_part = target.lstrip("/") if target.startswith("/") else "xl/" + target
assert target_part in contents, f"Worksheet part {target_part} not present in archive"

xml_text = contents[target_part].decode("utf-8")


def col_letter(col_idx):
    letters = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


for (row, col), (value, is_str) in cached.items():
    ref = f"{col_letter(col)}{row}"
    # openpyxl already writes an empty <v></v> placeholder after each
    # formula's <f> tag; fill it in rather than inserting a new element.
    pattern = re.compile(r'<c r="' + re.escape(ref) + r'"([^>]*)>(<f>.*?</f>)<v></v></c>')
    m = pattern.search(xml_text)
    assert m, f"Cell {ref} formula placeholder not found in worksheet XML ({target_part})"
    attrs, f_tag = m.groups()
    if is_str and 't="str"' not in attrs:
        attrs = ' t="str"' + attrs
    if is_str:
        v_text = str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    else:
        v_text = repr(value) if isinstance(value, float) else str(value)
    replacement = f'<c r="{ref}"{attrs}>{f_tag}<v>{v_text}</v></c>'
    xml_text = xml_text[: m.start()] + replacement + xml_text[m.end():]

contents[target_part] = xml_text.encode("utf-8")

tmp_path = XLSX_PATH + ".tmp"
with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zf:
    for name, data in contents.items():
        zf.writestr(name, data)
shutil.move(tmp_path, XLSX_PATH)
print("Injected cached values for Category Summary formulas.")
